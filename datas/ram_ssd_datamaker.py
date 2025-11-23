# 엑셀 파일을 다루기 위한 openpyxl 라이브러리입니다.
# 코드 실행 전 터미널(명령 프롬프트)에 'pip install openpyxl'을 입력하여 설치해주세요.

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
import os

# --- 기본 설정 ---
SHEET_NAMES = ['거래완료', '예약중', '판매중']
COMMON_HEADERS = [
    '판매가', '다나와 검색', '신제품 판매가', '좋아요 수', '채팅 수', '조회수',
    '미개봉 여부', '소유권 이전횟수', 'AS기간', '사용빈도', '구매일',
    '사진 개수', '파일명(복사용)', '사진1', '사진2', '사진3', '사진4', '사진5'
]

# 다른 파일과 동일한 날짜 선택 목록 (사용자 수식에 맞춰 "년 전" 제외)
DATE_LIST = ["오늘"] + [f"{i}일 전" for i in range(1, 31)] + \
            [f"{i}개월 전" for i in range(1, 13)]

# --- 헬퍼 함수: 스타일 및 기본 설정 ---

def apply_final_formatting(sheet, headers):
    """헤더 스타일, 셀 가운데 정렬, 그리고 내용에 맞는 최적의 열 너비를 설정합니다."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    for col_idx, header_cell in enumerate(sheet[1], 1):
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = center_align

        # 내용 기반 열 너비 자동 조절
        max_length = 0
        column = get_column_letter(col_idx)

        # 헤더 길이 계산
        if header_cell.value:
            max_length = len(str(header_cell.value))

        # 데이터 셀 길이 계산 (최대 100행까지만 샘플링)
        for cell in sheet[column][1:101]: # 100행까지만 검사
            if cell.value:
                # 수식인 경우 제외 (길이 계산이 부정확함)
                if not isinstance(cell.value, str) or not cell.value.startswith('='):
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

        # 열 너비 설정 (기본 여유 공간 2 추가)
        adjusted_width = (max_length + 2)

        # 특정 열 최소 너비 보장
        if header_cell.value in ['판매글 게시일자', '다나와 검색', '파일명(복사용)', '인터페이스']:
             if adjusted_width < 20:
                adjusted_width = 20

        sheet.column_dimensions[column].width = adjusted_width

    # 모든 데이터 셀 가운데 정렬
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = center_align

def apply_common_validations(sheet, headers):
    """모든 시트에 공통으로 적용되는 데이터 유효성 검사를 설정합니다."""

    # 드롭다운 목록이 저장된 숨겨진 시트 참조
    validation_sheet_name = "_ValidationLists"

    # 미개봉 여부 (O/X)
    if '미개봉 여부' in headers:
        unopened_col_letter = get_column_letter(headers.index('미개봉 여부') + 1)
        dv_unopened = DataValidation(type="list", formula1=f'={validation_sheet_name}!$A$1:$A$2', allow_blank=True)
        sheet.add_data_validation(dv_unopened)
        dv_unopened.add(f'{unopened_col_letter}2:{unopened_col_letter}1000')

    # 소유권 이전횟수 (0~5회, 모름)
    if '소유권 이전횟수' in headers:
        owner_col_letter = get_column_letter(headers.index('소유권 이전횟수') + 1)
        dv_owner = DataValidation(type="list", formula1=f'={validation_sheet_name}!$B$1:$B$7', allow_blank=True)
        sheet.add_data_validation(dv_owner)
        dv_owner.add(f'{owner_col_letter}2:{owner_col_letter}1000')

    # AS기간 (모름, 종료, 날짜)
    if 'AS기간' in headers:
        as_col_letter = get_column_letter(headers.index('AS기간') + 1)
        dv_as = DataValidation(type="list", formula1=f'={validation_sheet_name}!$C$1:$C$2', allow_blank=True)
        sheet.add_data_validation(dv_as)
        dv_as.add(f'{as_col_letter}2:{as_col_letter}1000')

    # 사용빈도 (모름, 주 0~10시간 등)
    if '사용빈도' in headers:
        usage_col_letter = get_column_letter(headers.index('사용빈도') + 1)
        dv_usage = DataValidation(type="list", formula1=f'={validation_sheet_name}!$D$1:$D$6', allow_blank=True)
        sheet.add_data_validation(dv_usage)
        dv_usage.add(f'{usage_col_letter}2:{usage_col_letter}1000')

    # 구매일 (모름, 약 1년 전 등)
    if '구매일' in headers:
        purchase_date_col_letter = get_column_letter(headers.index('구매일') + 1)
        dv_purchase = DataValidation(type="list", formula1=f'={validation_sheet_name}!$A$10:$A$17', allow_blank=True)
        sheet.add_data_validation(dv_purchase)
        dv_purchase.add(f'{purchase_date_col_letter}2:{purchase_date_col_letter}1000')

    # 사진 개수 (1~5)
    if '사진 개수' in headers:
        photo_count_col_letter = get_column_letter(headers.index('사진 개수') + 1)
        dv_photo_count = DataValidation(type="list", formula1=f'={validation_sheet_name}!$E$1:$E$5', allow_blank=True)
        sheet.add_data_validation(dv_photo_count)
        dv_photo_count.add(f'{photo_count_col_letter}2:{photo_count_col_letter}1000')

def create_validation_sheet(wb):
    """데이터 유효성 검사에 사용될 목록이 포함된 숨겨진 시트를 생성합니다."""
    if "_ValidationLists" not in wb.sheetnames:
        validation_sheet = wb.create_sheet(title="_ValidationLists")
        validation_sheet.sheet_state = 'hidden' # 시트 숨기기

        # 공통 목록
        common_lists = {
            "A": ["O", "X"], # 미개봉 여부
            "B": ["모름", "0", "1", "2", "3", "4", "5"], # 소유권 이전횟수
            "C": ["모름", "종료"], # AS기간
            "D": ["모름", "1주 0~10시간", "1주 10~20시간", "1주 20~30시간", "1주 30~40시간", "1주 40시간 이상"], # 사용빈도
            "E": ["1", "2", "3", "4", "5"] # 사진 개수
        }

        # 구매일 목록 (A열 10행부터)
        purchase_date_list = ["모름", "약 1개월 전", "약 3개월 전", "약 6개월 전", "약 1년 전", "약 2년 전", "약 3년 전", "약 4년 전 이상"]
        for i, val in enumerate(purchase_date_list, 10):
            validation_sheet[f'A{i}'] = val

        for col, values in common_lists.items():
            for row, value in enumerate(values, 1):
                validation_sheet[f'{col}{row}'] = value

        return validation_sheet
    else:
        return wb["_ValidationLists"]

# --- SSD 엑셀 파일 생성 ---

def create_ssd_excel(wb):
    """SSD.xlsx 파일을 생성하고 시트를 설정합니다. ⭐ 인터페이스 컬럼 추가"""
    # ⭐ 인터페이스 컬럼 추가 (25번째)
    headers = [
        '판매글 게시일자', '날짜 선택', '제조사', '시리즈/모델명', '폼팩터', '용량',
    ] + COMMON_HEADERS + ['인터페이스']

    validation_sheet = create_validation_sheet(wb)
    validation_sheet_name = "_ValidationLists"

    # SSD 관련 목록 (F열부터 시작)
    ssd_lists = {
        "F": ["삼성전자", "SK하이닉스", "Western Digital", "Micron Crucial", "Seagate", "기타"],
        "G": [
            "990 PRO", "980 PRO", "970 EVO Plus", "870 EVO", "PM9A1",
            "Platinum P41", "Gold P31",
            "WD BLACK SN850X", "WD Blue SN580", "WD Green SN350",
            "P5 Plus", "MX500", "BX500",
            "FireCuda 530", "BarraCuda Q5",
            "기타"
        ],
        "H": ["M.2 (2280)", "M.2 (2242)", "M.2 NVMe", "M.2 SATA", "2.5인치", "Portable", "기타"],
        "I": ["128GB", "250GB", "500GB", "1TB", "2TB", "4TB", "기타"],
        # ⭐ 인터페이스 목록 추가 (P열)
        "P": ["NVMe", "SATA"]
    }

    for col, values in ssd_lists.items():
        for row, value in enumerate(values, 1):
            validation_sheet[f'{col}{row}'] = value

    # 날짜 목록 (J열)
    for row, value in enumerate(DATE_LIST, 1):
        validation_sheet[f'J{row}'] = value

    for sheet_name in SHEET_NAMES:
        sheet = wb.create_sheet(title=sheet_name)
        sheet.append(headers)

        # 데이터 유효성 검사 (드롭다운)
        dv_date_select = DataValidation(type="list", formula1=f'={validation_sheet_name}!$J$1:$J${len(DATE_LIST)}', allow_blank=True)
        sheet.add_data_validation(dv_date_select)
        dv_date_select.add('B2:B1000')

        dv_manufacturer = DataValidation(type="list", formula1=f'={validation_sheet_name}!$F$1:$F${len(ssd_lists["F"])}', allow_blank=True)
        sheet.add_data_validation(dv_manufacturer)
        dv_manufacturer.add('C2:C1000')

        dv_series = DataValidation(type="list", formula1=f'={validation_sheet_name}!$G$1:$G${len(ssd_lists["G"])}', allow_blank=True)
        sheet.add_data_validation(dv_series)
        dv_series.add('D2:D1000')

        dv_form_factor = DataValidation(type="list", formula1=f'={validation_sheet_name}!$H$1:$H${len(ssd_lists["H"])}', allow_blank=True)
        sheet.add_data_validation(dv_form_factor)
        dv_form_factor.add('E2:E1000')

        dv_capacity = DataValidation(type="list", formula1=f'={validation_sheet_name}!$I$1:$I${len(ssd_lists["I"])}', allow_blank=True)
        sheet.add_data_validation(dv_capacity)
        dv_capacity.add('F2:F1000')

        # ⭐ 인터페이스 유효성 검사 추가
        interface_col_idx = headers.index('인터페이스') + 1
        dv_interface = DataValidation(type="list", formula1=f'={validation_sheet_name}!$P$1:$P${len(ssd_lists["P"])}', allow_blank=True)
        dv_interface.errorStyle = 'warning'
        dv_interface.promptTitle = '입력 안내'
        dv_interface.prompt = '목록에서 인터페이스를 선택하세요.'
        dv_interface.showInputMessage = True
        dv_interface.add(f'{get_column_letter(interface_col_idx)}2:{get_column_letter(interface_col_idx)}1000')
        sheet.add_data_validation(dv_interface)

        # 공통 유효성 검사 적용
        apply_common_validations(sheet, headers)

        # 수식 및 기본값 설정
        danawa_col_idx = headers.index('다나와 검색') + 1
        copy_filename_col_idx = headers.index('파일명(복사용)') + 1
        photo_count_col_idx = headers.index('사진 개수') + 1

        # 기본값 설정용 열 인덱스
        unopened_col_idx = headers.index('미개봉 여부') + 1
        owner_col_idx = headers.index('소유권 이전횟수') + 1
        as_col_idx = headers.index('AS기간') + 1
        usage_col_idx = headers.index('사용빈도') + 1
        purchase_date_col_idx = headers.index('구매일') + 1

        for row in range(2, 1001): # 1000행까지 수식 적용

            # 1. B열(날짜선택) 선택 시 A열(판매글 게시일자) 자동 계산
            # 기준 날짜는 오늘 날짜로 설정
            today_str = date.today().strftime("%Y-%m-%d")
            date_formula = f'=IF(B{row}="오늘",DATEVALUE("{today_str}"), IF(ISNUMBER(FIND("일 전", B{row})), DATEVALUE("{today_str}") - VALUE(SUBSTITUTE(B{row},"일 전","")), IF(ISNUMBER(FIND("개월 전", B{row})), EDATE(DATEVALUE("{today_str}"), -VALUE(SUBSTITUTE(B{row},"개월 전",""))), "")))'
            cell = sheet.cell(row=row, column=1, value=date_formula)
            cell.number_format = 'YYYY-MM-DD' # 표시 형식 설정

            # 2. B열(날짜선택) 선택 시 기본값 자동 입력
            default_value_formula_unopened = f'=IF(B{row}<>"", "X", "")'
            sheet.cell(row=row, column=unopened_col_idx, value=default_value_formula_unopened)

            default_value_formula_owner = f'=IF(B{row}<>"", "모름", "")'
            sheet.cell(row=row, column=owner_col_idx, value=default_value_formula_owner)

            default_value_formula_as = f'=IF(B{row}<>"", "모름", "")'
            sheet.cell(row=row, column=as_col_idx, value=default_value_formula_as)

            default_value_formula_usage = f'=IF(B{row}<>"", "모름", "")'
            sheet.cell(row=row, column=usage_col_idx, value=default_value_formula_usage)

            default_value_formula_purchase = f'=IF(B{row}<>"", "모름", "")'
            sheet.cell(row=row, column=purchase_date_col_idx, value=default_value_formula_purchase)

            default_value_formula_photo_count = f'=IF(B{row}<>"", "1", "")'
            sheet.cell(row=row, column=photo_count_col_idx, value=default_value_formula_photo_count)

            # 3. C열(제조사) 등 입력 시 '다나와 검색' 및 '파일명' 자동 완성
            search_query = f'C{row}&" "&D{row}&" "&E{row}&" "&F{row}'
            base_filename_formula = f'SUBSTITUTE(TRIM({search_query}), " ", "_")'

            # 다나와 검색 (C열이 비어있지 않으면 실행)
            danawa_formula = f'=IF(C{row}<>"", HYPERLINK("https://search.danawa.com/dsearch.php?query=" & {search_query}, "다나와 가격 확인"), "")'
            sheet.cell(row=row, column=danawa_col_idx, value=danawa_formula)

            # 파일명 (C열이 비어있지 않으면 실행)
            copy_filename_formula = f'=IF(C{row}<>"", {base_filename_formula} & "_" & TEXT(A{row}, "YYYYMMDD") & "_", "")'
            sheet.cell(row=row, column=copy_filename_col_idx, value=copy_filename_formula)

            # 사진 링크 (파일명이 존재하면 실행)
            for i in range(1, 6):
                photo_col_idx = headers.index(f'사진{i}') + 1
                photo_formula = f'=IF(AND(C{row}<>"", IFERROR(VALUE({get_column_letter(photo_count_col_idx)}{row}), 0)>={i}), HYPERLINK("images\\\\" & {base_filename_formula} & "_" & TEXT(A{row}, "YYYYMMDD") & "_{i}.png", {base_filename_formula} & "_" & TEXT(A{row}, "YYYYMMDD") & "_{i}.png"), "")'
                sheet.cell(row=row, column=photo_col_idx, value=photo_formula)

        apply_final_formatting(sheet, headers)


# --- RAM 엑셀 파일 생성 ---

def create_ram_excel(wb):
    """RAM.xlsx 파일을 생성하고 시트를 설정합니다. ⭐ 판매 개수와 판매가(개당) 포함"""
    # ⭐ RAM 전용 헤더: 판매 개수와 판매가(개당) 추가
    headers = [
        '판매글 게시일자', '날짜 선택', '제조사', '메모리 규격', '시리즈/모델명', '클럭 (MHz)', '용량 (GB)',
        '판매 개수', '판매가(개당)', '다나와 검색', '신제품 판매가', '좋아요 수', '채팅 수', '조회수',
        '미개봉 여부', '소유권 이전횟수', 'AS기간', '사용빈도', '구매일',
        '사진 개수', '파일명(복사용)', '사진1', '사진2', '사진3', '사진4', '사진5'
    ]

    validation_sheet = create_validation_sheet(wb)
    validation_sheet_name = "_ValidationLists"

    # RAM 관련 목록 (K열부터 시작)
    ram_lists = {
        "K": ["삼성전자", "SK하이닉스", "Corsair", "G.Skill", "Kingston", "TeamGroup", "Micron Crucial", "기타"],
        "L": ["DDR5", "DDR4", "DDR3", "SO-DIMM DDR5", "SO-DIMM DDR4", "기타"],
        "M": [
            "기본 (시금치)",
            "Vengeance", "Dominator",
            "Trident Z", "Ripjaws",
            "FURY Renegade", "FURY Beast",
            "T-Force Delta", "T-Create",
            "기타"
        ],
        "N": ["8000", "7200", "6400", "6000", "5600", "5200", "4800", "3600", "3200", "2666", "2400", "1600", "기타"],
        "O": ["8GB", "16GB", "32GB (2x16GB)", "32GB", "64GB (2x32GB)", "기타"]
    }

    for col, values in ram_lists.items():
        for row, value in enumerate(values, 1):
            validation_sheet[f'{col}{row}'] = value

    # 날짜 목록 (J열 - SSD와 공유)
    if validation_sheet['J1'].value is None:
        for row, value in enumerate(DATE_LIST, 1):
            validation_sheet[f'J{row}'] = value

    for sheet_name in SHEET_NAMES:
        sheet = wb.create_sheet(title=sheet_name)
        sheet.append(headers)

        # 데이터 유효성 검사 (드롭다운)
        dv_date_select = DataValidation(type="list", formula1=f'={validation_sheet_name}!$J$1:$J${len(DATE_LIST)}', allow_blank=True)
        sheet.add_data_validation(dv_date_select)
        dv_date_select.add('B2:B1000')

        dv_manufacturer = DataValidation(type="list", formula1=f'={validation_sheet_name}!$K$1:$K${len(ram_lists["K"])}', allow_blank=True)
        sheet.add_data_validation(dv_manufacturer)
        dv_manufacturer.add('C2:C1000')

        dv_spec = DataValidation(type="list", formula1=f'={validation_sheet_name}!$L$1:$L${len(ram_lists["L"])}', allow_blank=True)
        sheet.add_data_validation(dv_spec)
        dv_spec.add('D2:D1000')

        dv_model = DataValidation(type="list", formula1=f'={validation_sheet_name}!$M$1:$M${len(ram_lists["M"])}', allow_blank=True)
        sheet.add_data_validation(dv_model)
        dv_model.add('E2:E1000')

        dv_clock = DataValidation(type="list", formula1=f'={validation_sheet_name}!$N$1:$N${len(ram_lists["N"])}', allow_blank=True)
        sheet.add_data_validation(dv_clock)
        dv_clock.add('F2:F1000')

        dv_capacity = DataValidation(type="list", formula1=f'={validation_sheet_name}!$O$1:$O${len(ram_lists["O"])}', allow_blank=True)
        sheet.add_data_validation(dv_capacity)
        dv_capacity.add('G2:G1000')

        # 공통 유효성 검사 적용
        apply_common_validations(sheet, headers)

        # 수식 및 기본값 설정
        danawa_col_idx = headers.index('다나와 검색') + 1
        copy_filename_col_idx = headers.index('파일명(복사용)') + 1
        photo_count_col_idx = headers.index('사진 개수') + 1

        # 기본값 설정용 열 인덱스
        unopened_col_idx = headers.index('미개봉 여부') + 1
        owner_col_idx = headers.index('소유권 이전횟수') + 1
        as_col_idx = headers.index('AS기간') + 1
        usage_col_idx = headers.index('사용빈도') + 1
        purchase_date_col_idx = headers.index('구매일') + 1

        for row in range(2, 1001): # 1000행까지 수식 적용

            # 1. B열(날짜선택) 선택 시 A열(판매글 게시일자) 자동 계산
            today_str = date.today().strftime("%Y-%m-%d")
            date_formula = f'=IF(B{row}="오늘",DATEVALUE("{today_str}"), IF(ISNUMBER(FIND("일 전", B{row})), DATEVALUE("{today_str}") - VALUE(SUBSTITUTE(B{row},"일 전","")), IF(ISNUMBER(FIND("개월 전", B{row})), EDATE(DATEVALUE("{today_str}"), -VALUE(SUBSTITUTE(B{row},"개월 전",""))), "")))'
            cell = sheet.cell(row=row, column=1, value=date_formula)
            cell.number_format = 'YYYY-MM-DD' # 표시 형식 설정

            # 2. B열(날짜선택) 선택 시 기본값 자동 입력
            default_value_formula_unopened = f'=IF(B{row}<>"", "X", "")'
            sheet.cell(row=row, column=unopened_col_idx, value=default_value_formula_unopened)

            default_value_formula_owner = f'=IF(B{row}<>"", "모름", "")'
            sheet.cell(row=row, column=owner_col_idx, value=default_value_formula_owner)

            default_value_formula_as = f'=IF(B{row}<>"", "모름", "")'
            sheet.cell(row=row, column=as_col_idx, value=default_value_formula_as)

            default_value_formula_usage = f'=IF(B{row}<>"", "모름", "")'
            sheet.cell(row=row, column=usage_col_idx, value=default_value_formula_usage)

            default_value_formula_purchase = f'=IF(B{row}<>"", "모름", "")'
            sheet.cell(row=row, column=purchase_date_col_idx, value=default_value_formula_purchase)

            default_value_formula_photo_count = f'=IF(B{row}<>"", "1", "")'
            sheet.cell(row=row, column=photo_count_col_idx, value=default_value_formula_photo_count)

            # 3. C열(제조사) 등 입력 시 '다나와 검색' 및 '파일명' 자동 완성
            search_query = f'C{row}&" "&D{row}&" "&E{row}&" "&F{row}&" "&G{row}'
            base_filename_formula = f'SUBSTITUTE(TRIM({search_query}), " ", "_")'

            # 다나와 검색 (C열이 비어있지 않으면 실행)
            danawa_formula = f'=IF(C{row}<>"", HYPERLINK("https://search.danawa.com/dsearch.php?query=" & {search_query}, "다나와 가격 확인"), "")'
            sheet.cell(row=row, column=danawa_col_idx, value=danawa_formula)

            # 파일명 (C열이 비어있지 않으면 실행)
            copy_filename_formula = f'=IF(C{row}<>"", {base_filename_formula} & "_" & TEXT(A{row}, "YYYYMMDD") & "_", "")'
            sheet.cell(row=row, column=copy_filename_col_idx, value=copy_filename_formula)

            # 사진 링크 (파일명이 존재하면 실행)
            for i in range(1, 6):
                photo_col_idx = headers.index(f'사진{i}') + 1
                photo_formula = f'=IF(AND(C{row}<>"", IFERROR(VALUE({get_column_letter(photo_count_col_idx)}{row}), 0)>={i}), HYPERLINK("images\\\\" & {base_filename_formula} & "_" & TEXT(A{row}, "YYYYMMDD") & "_{i}.png", {base_filename_formula} & "_" & TEXT(A{row}, "YYYYMMDD") & "_{i}.png"), "")'
                sheet.cell(row=row, column=photo_col_idx, value=photo_formula)

        apply_final_formatting(sheet, headers)


# --- 메인 실행 ---
if __name__ == "__main__":
    # 스크립트가 실행되는 디렉토리로 작업 경로를 변경합니다.
    # 이렇게 하면 항상 스크립트 파일 위치에 엑셀 파일이 생성됩니다.
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)

    try:
        print("SSD.xlsx 파일 생성을 시작합니다... (인터페이스 컬럼 포함)")
        ssd_wb = Workbook()
        # 기본 'Sheet' 시트 제거
        if "Sheet" in ssd_wb.sheetnames:
            ssd_wb.remove(ssd_wb["Sheet"])
        create_ssd_excel(ssd_wb)
        ssd_wb.save('SSD.xlsx')
        print("✅ SSD.xlsx 파일 생성이 완료되었습니다.")

        print("RAM.xlsx 파일 생성을 시작합니다... (판매 개수, 판매가(개당) 컬럼 포함)")
        ram_wb = Workbook()
        # 기본 'Sheet' 시트 제거
        if "Sheet" in ram_wb.sheetnames:
            ram_wb.remove(ram_wb["Sheet"])
        create_ram_excel(ram_wb)
        ram_wb.save('RAM.xlsx')
        print("✅ RAM.xlsx 파일 생성이 완료되었습니다.")

        print("\n🎉 모든 작업이 완료되었습니다!")
        print("📍 생성 위치: 스크립트 파일과 동일한 디렉토리")
        print("\n✅ 추가된 컬럼:")
        print("   - SSD.xlsx: 인터페이스 (NVMe/SATA)")
        print("   - RAM.xlsx: 판매 개수, 판매가(개당) (총 26개 컬럼)")
        print("\n📊 최종 컬럼 수:")
        print("   - RAM.xlsx: 26개 (업로드된 파일과 동일)")
        print("   - SSD.xlsx: 25개 (인터페이스 추가)")
    except Exception as e:
        print(f"❌ 오류가 발생했습니다: {e}")
        print("열려있는 엑셀 파일을 모두 닫고 다시 시도해보세요.")
        print("openpyxl 라이브러리가 설치되었는지 확인해주세요. (터미널에서 pip install openpyxl 실행)")