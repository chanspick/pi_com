# 엑셀 파일을 다루기 위한 openpyxl 라이브러리입니다.
# 코드 실행 전 터미널(명령 프롬프트)에 'pip install openpyxl'을 입력하여 설치해주세요.

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
import os

# --- 기본 설정 ---
SHEET_NAMES = ['거래완료', '예약중', '판매중']
COMMON_HEADERS = [
    '판매가', '다나와 검색', '신제품 판매가', '좋아요 수', '채팅 수', '조회수', 
    '미개봉 여부', '소유권 이전횟수', 'AS기간', '사용빈도', '구매일', 
    '사진 개수', '파일명(복사용)', '사진1', '사진2', '사진3', '사진4', '사진5'
]

# --- 헬퍼 함수: 스타일 및 기본 설정 ---

def apply_final_formatting(sheet, headers):
    """헤더 스타일, 셀 가운데 정렬, 그리고 내용에 맞는 최적의 열 너비를 설정합니다."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    for col_idx, header_cell in enumerate(sheet[1], 1):
        header_cell.font = header_font
        header_cell.fill = header_fill
        for row_idx in range(1, 1002):
            sheet.cell(row=row_idx, column=col_idx).alignment = center_align

    predefined_widths = {
        '브랜드': 20, '제조사': 25, '시리즈': 22, '세부 모델명': 25, '모델명': 22,
        '파일명(복사용)': 60, '다나와 검색': 18, '판매가': 15, '신제품 판매가': 15,
        'AS기간': 15, '구매일': 20, '판매글 게시일자': 20, '날짜 선택': 15, '사진 개수': 12,
        '사용빈도': 20, '소유권 이전횟수': 15, '소켓': 18, 'TDP (W)': 12,
        '메모리 타입': 15, '폼팩터': 15
    }
    for i in range(1, 6):
        predefined_widths[f'사진{i}'] = 65


    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        if header in predefined_widths:
            sheet.column_dimensions[col_letter].width = predefined_widths[header]
        else:
            sheet.column_dimensions[col_letter].width = len(header) + 5


def create_validation_lists_sheet(workbook):
    """데이터 유효성 검사에 사용할 목록을 별도의 숨겨진 시트에 만듭니다."""
    if '_ValidationLists' in workbook.sheetnames:
        return workbook['_ValidationLists']

    val_sheet = workbook.create_sheet(title="_ValidationLists")

    date_options = ['오늘'] + [f'{i}일 전' for i in range(1, 31)] + [f'{i}개월 전' for i in range(1, 13)]
    usage_options = ['모름'] + [f'1주 {i*10}~{(i+1)*10}시간' for i in range(14)]

    lists = {
        'AS_Period_Options': ['모름', '종료'],
        'Ownership_Transfer_Count': ['모름', 0, 1, 2, 3, 4, 5],
        'Usage_Frequency': usage_options,
        'Photo_Count': [1, 2, 3, 4, 5],
        'Date_Selection': date_options,
        'CPU_Brand': ['Intel Core', 'AMD Ryzen'],
        'Intel_Core_Series': ['i9', 'i7', 'i5', 'i3'],
        'AMD_Ryzen_Series': ['9', '7', '5', '3'],
        'Intel_Core_i9_Models': [14900, 13900, 12900, 11900, 10900],
        'Intel_Core_i7_Models': [14700, 13700, 12700, 11700, 10700],
        'Intel_Core_i5_Models': [14600, 13600, 13500, 13400, 12600, 12500, 12400, 11600, 11400, 10400],
        'Intel_Core_i3_Models': [13100, 12100, 10100],
        'AMD_Ryzen_9_Models': [7950, 7900, 5950, 5900],
        'AMD_Ryzen_7_Models': [7800, 7700, 5800, 5700],
        'AMD_Ryzen_5_Models': [7600, 7500, 5600, 5500],
        'AMD_Ryzen_3_Models': [5300, 4300, 4100],
        'Intel_Core_Suffix': ['', 'K', 'F', 'KF', 'KS', 'T', 'H', 'HX', 'P', 'U'],
        'AMD_Ryzen_Suffix': ['', 'X', 'X3D', 'G', 'GE', 'H', 'HS', 'HX', 'U'],

        # ⭐ CPU 소켓 추가
        'CPU_Socket': ['AM4', 'AM5', 'LGA1151', 'LGA1200', 'LGA1700', 'LGA1851', 'TR4', 'sTRX4'],

        'GPU_Brand': ['NVIDIA GeForce', 'AMD Radeon'],
        'GPU_Manufacturer': ['ASUS', 'Gigabyte', 'MSI', 'Zotac', 'EVGA', 'Sapphire', 'PowerColor', 'XFX', 'Palit', 'Gainward', 'Inno3D', 'Colorful', 'PNY', 'Manli', 'Leadtek', 'Biostar', '갤럭시 (GALAX)'],
        'NVIDIA_GeForce_Series': ['RTX', 'GTX'],
        'AMD_Radeon_Series': ['RX'],
        'NVIDIA_GeForce_RTX_Models': [5090, 5080, 5070, 5060, 4090, 4080, 4070, 4060, 3090, 3080, 3070, 3060, 3050, 2080, 2070, 2060],
        'NVIDIA_GeForce_GTX_Models': [1660, 1650, 1080, 1070, 1060, 1050],
        'AMD_Radeon_RX_Models': [8900, 8800, 8700, 7900, 7800, 7700, 7600, 6950, 6900, 6800, 6700, 6600, 6500],
        'NVIDIA_GeForce_Suffix': ['', 'Ti', 'Super'],
        'AMD_Radeon_Suffix': ['', 'XT', 'XTX'],
        'ASUS_GPU_Detailed_Models': ['ROG Strix', 'TUF Gaming', 'Dual', 'KO', 'ProArt', 'Phoenix'],
        'Gigabyte_GPU_Detailed_Models': ['Aorus Master', 'Aorus Elite', 'Gaming OC', 'Windforce', 'Eagle', 'Vision'],
        'MSI_GPU_Detailed_Models': ['Suprim', 'Gaming Trio', 'Ventus', 'Sea Hawk', 'Aero ITX'],
        'Zotac_GPU_Detailed_Models': ['AMP Extreme', 'Trinity', 'Twin Edge'],
        '갤럭시_(GALAX)_GPU_Detailed_Models': ['SG', 'EX Gamer', 'HOF', 'BOY'],
        'Sapphire_GPU_Detailed_Models': ['Nitro+', 'Pulse', 'Toxic'],
        'PowerColor_GPU_Detailed_Models': ['Red Devil', 'Hellhound', 'Fighter'],
        'XFX_GPU_Detailed_Models': ['MERC', 'QICK', 'SWFT'],
        'GPU_Memory_Capacity': ['', '32GB', '24GB', '20GB', '16GB', '12GB', '10GB', '8GB', '6GB', '4GB'],

        # ⭐ GPU TDP 추가 (W 단위, 숫자만)
        'GPU_TDP': ['', '450', '355', '350', '335', '320', '300', '290', '285', '263', '250', '245', '230', '225', '220', '200', '180', '170', '165', '160', '150', '132', '130', '115', '107'],

        'Unopened_Status': ['O', 'X'],
        'Mainboard_Manufacturer': ['ASUS', 'Gigabyte', 'MSI', 'ASRock', 'Biostar', 'Colorful'],
        'ASUS_Series': ['ROG', 'TUF Gaming', 'Prime', 'ProArt'],
        'Gigabyte_Series': ['Aorus', 'Gaming X', 'Aero', 'UD'],
        'MSI_Series': ['MEG', 'MPG', 'MAG', 'PRO'],
        'ASRock_Series': ['Taichi', 'Phantom Gaming', 'Steel Legend', 'Pro RS'],
        'Biostar_Series': ['Racing', 'Valkyrie'],
        'Colorful_Series': ['iGame', 'CVN', 'Battle-Axe'],
        'ASUS_ROG_Models': ['Maximus', 'Strix', 'Hero', 'Apex', 'Formula'],
        'ASUS_TUF_Gaming_Models': ['Plus', 'Pro'],
        'ASUS_Prime_Models': ['-P', '-A', '-K'],
        'ASUS_ProArt_Models': ['Creator'],
        'Gigabyte_Aorus_Models': ['Master', 'Elite', 'Pro', 'Xtreme', 'Tachyon'],
        'Gigabyte_Gaming_X_Models': ['Gaming X'],
        'Gigabyte_Aero_Models': ['Aero G'],
        'Gigabyte_UD_Models': ['UD'],
        'MSI_MEG_Models': ['Godlike', 'Ace', 'Unify'],
        'MSI_MPG_Models': ['Carbon', 'Edge', 'Force', 'Velox'],
        'MSI_MAG_Models': ['Tomahawk', 'Mortar', 'Torpedo', 'Bazooka'],
        'MSI_PRO_Models': ['-A', '-P'],
        'ASRock_Taichi_Models': ['Taichi'],
        'ASRock_Phantom_Gaming_Models': ['PG', 'Velocita', 'Riptide', 'Lightning'],
        'ASRock_Steel_Legend_Models': ['Steel Legend'],
        'ASRock_Pro_RS_Models': ['Pro RS'],
        'Biostar_Racing_Models': ['Racing'],
        'Biostar_Valkyrie_Models': ['Valkyrie'],
        'Colorful_iGame_Models': ['Ultra', 'Vulcan'],
        'Colorful_CVN_Models': ['CVN'],
        'Colorful_Battle_Axe_Models': ['Battle-Axe'],

        'Mainboard_Chipset': ['Z790', 'B760', 'H770', 'Z690', 'B660', 'H610', 'X670E', 'X670', 'B650E', 'B650', 'A620', 'X570', 'B550', 'A520'],
        'Mainboard_Suffix': ['', 'Pro', 'Plus', 'Wi-Fi', 'Gaming', 'D4', 'D5', 'M-ATX', 'ITX', '-E', '-F', '-I', 'AC', 'AX', 'Lightning', 'Edge', 'Tomahawk', 'Carbon'],

        # ⭐ 메인보드 소켓, 메모리 타입, 폼팩터 추가
        'Mainboard_Socket': ['AM4', 'AM5', 'LGA1151', 'LGA1200', 'LGA1700', 'LGA1851', 'TR4', 'sTRX4', 'AM3+'],
        'Mainboard_Memory_Type': ['DDR3', 'DDR4', 'DDR5'],
        'Mainboard_Form_Factor': ['ATX', 'mATX', 'Mini-ITX', 'E-ATX']
    }

    col_idx = 1
    for name, items in lists.items():
        val_sheet.cell(row=1, column=col_idx, value=name)
        for row_idx, item in enumerate(items, 1):
            val_sheet.cell(row=row_idx, column=col_idx, value=item)
        range_string = f"'{val_sheet.title}'!${get_column_letter(col_idx)}$1:${get_column_letter(col_idx)}${len(items)}"
        defined_name = name.replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")
        named_range = openpyxl.workbook.defined_name.DefinedName(defined_name, attr_text=range_string)
        workbook.defined_names[defined_name] = named_range
        col_idx += 1

    val_sheet.sheet_state = 'hidden'
    return val_sheet


def apply_common_validations(sheet, headers):
    """여러 시트에 공통적으로 적용될 데이터 유효성 검사 및 기본값을 설정합니다."""
    header_map = {header: i + 1 for i, header in enumerate(headers)}

    price_cols = [header_map.get('판매가'), header_map.get('신제품 판매가')]
    for col_idx in price_cols:
        if col_idx:
            for row in range(2, 1001):
                sheet.cell(row=row, column=col_idx).number_format = '#,##0'

    dv_unopened = DataValidation(type="list", formula1='=Unopened_Status', allow_blank=True)
    dv_unopened.errorStyle = 'warning'
    dv_unopened.promptTitle = '입력 안내'
    dv_unopened.prompt = 'O 또는 X를 선택하세요.'
    dv_unopened.showInputMessage = True
    if header_map.get('미개봉 여부'):
        col_letter = get_column_letter(header_map['미개봉 여부'])
        dv_unopened.add(f'{col_letter}2:{col_letter}1000')
        sheet.add_data_validation(dv_unopened)

    dv_ownership = DataValidation(type="list", formula1='=Ownership_Transfer_Count', allow_blank=True)
    dv_ownership.errorStyle = 'warning'
    dv_ownership.promptTitle = '입력 안내'
    dv_ownership.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
    dv_ownership.showInputMessage = True
    if header_map.get('소유권 이전횟수'):
        col_letter = get_column_letter(header_map['소유권 이전횟수'])
        dv_ownership.add(f'{col_letter}2:{col_letter}1000')
        sheet.add_data_validation(dv_ownership)

    dv_as_period = DataValidation(type="list", formula1='=AS_Period_Options', allow_blank=True)
    dv_as_period.errorStyle = 'warning'
    dv_as_period.promptTitle = '입력 안내'
    dv_as_period.prompt = '목록에서 선택하거나 날짜를 입력할 수 있습니다.'
    dv_as_period.showInputMessage = True
    if header_map.get('AS기간'):
        col_letter = get_column_letter(header_map['AS기간'])
        dv_as_period.add(f'{col_letter}2:{col_letter}1000')
        sheet.add_data_validation(dv_as_period)

    dv_usage = DataValidation(type="list", formula1='=Usage_Frequency', allow_blank=True)
    dv_usage.errorStyle = 'warning'
    dv_usage.promptTitle = '입력 안내'
    dv_usage.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
    dv_usage.showInputMessage = True
    if header_map.get('사용빈도'):
        col_letter = get_column_letter(header_map['사용빈도'])
        dv_usage.add(f'{col_letter}2:{col_letter}1000')
        sheet.add_data_validation(dv_usage)


def create_cpu_excel():
    """CPU.xlsx 파일을 생성합니다. ⭐ 소켓 컬럼 추가"""
    wb = Workbook()
    wb.remove(wb.active)
    create_validation_lists_sheet(wb)

    # ⭐ 소켓 컬럼 추가 (25번째)
    headers = ['판매글 게시일자', '날짜 선택', '브랜드', '시리즈', '모델 번호', '접미사'] + COMMON_HEADERS + ['소켓']
    today_str = date.today().strftime("%Y-%m-%d")

    for sheet_name in SHEET_NAMES:
        sheet = wb.create_sheet(sheet_name)
        sheet.append(headers)

        dv_date = DataValidation(type="list", formula1='=Date_Selection', allow_blank=True)
        dv_date.errorStyle = 'warning'
        dv_date.promptTitle = '입력 안내'
        dv_date.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_date.showInputMessage = True
        dv_date.add('B2:B1000')
        sheet.add_data_validation(dv_date)

        dv_brand = DataValidation(type="list", formula1='=CPU_Brand', allow_blank=True)
        dv_brand.errorStyle = 'warning'
        dv_brand.promptTitle = '입력 안내'
        dv_brand.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_brand.showInputMessage = True
        dv_brand.add('C2:C1000')
        sheet.add_data_validation(dv_brand)

        dv_series = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(C2," ","_")&"_Series")', allow_blank=True)
        dv_series.errorStyle = 'warning'
        dv_series.promptTitle = '입력 안내'
        dv_series.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_series.showInputMessage = True
        dv_series.add('D2:D1000')
        sheet.add_data_validation(dv_series)

        dv_model = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(C2," ","_")&"_"&D2&"_Models")', allow_blank=True)
        dv_model.errorStyle = 'warning'
        dv_model.promptTitle = '입력 안내'
        dv_model.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_model.showInputMessage = True
        dv_model.add('E2:E1000')
        sheet.add_data_validation(dv_model)

        dv_suffix = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(C2," ","_")&"_Suffix")', allow_blank=True)
        dv_suffix.errorStyle = 'warning'
        dv_suffix.promptTitle = '입력 안내'
        dv_suffix.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_suffix.showInputMessage = True
        dv_suffix.add('F2:F1000')
        sheet.add_data_validation(dv_suffix)

        # ⭐ 소켓 유효성 검사 추가
        socket_col_idx = headers.index('소켓') + 1
        dv_socket = DataValidation(type="list", formula1='=CPU_Socket', allow_blank=True)
        dv_socket.errorStyle = 'warning'
        dv_socket.promptTitle = '입력 안내'
        dv_socket.prompt = '목록에서 소켓을 선택하세요.'
        dv_socket.showInputMessage = True
        dv_socket.add(f'{get_column_letter(socket_col_idx)}2:{get_column_letter(socket_col_idx)}1000')
        sheet.add_data_validation(dv_socket)

        photo_count_col_idx = headers.index('사진 개수') + 1
        copy_filename_col_idx = headers.index('파일명(복사용)') + 1
        danawa_col_idx = headers.index('다나와 검색') + 1

        dv_photo_count = DataValidation(type="list", formula1='=Photo_Count', allow_blank=True)
        dv_photo_count.errorStyle = 'warning'
        dv_photo_count.promptTitle = '입력 안내'
        dv_photo_count.prompt = '목록에서 사진 개수를 선택하세요.'
        dv_photo_count.showInputMessage = True
        dv_photo_count.add(f'{get_column_letter(photo_count_col_idx)}2:{get_column_letter(photo_count_col_idx)}1000')
        sheet.add_data_validation(dv_photo_count)

        for row in range(2, 1001):
            date_formula = f'=IF(B{row}="오늘",DATEVALUE("{today_str}"), IF(ISNUMBER(FIND("일 전", B{row})), DATEVALUE("{today_str}") - VALUE(SUBSTITUTE(B{row},"일 전","")), IF(ISNUMBER(FIND("개월 전", B{row})), EDATE(DATEVALUE("{today_str}"), -VALUE(SUBSTITUTE(B{row},"개월 전",""))), "")))'
            sheet.cell(row=row, column=1, value=date_formula)

            search_query = 'C{0}&" "&D{0}&" "&E{0}&F{0}'.format(row)
            base_filename_formula = 'SUBSTITUTE(C{0}," ","_")&"_"&D{0}&"_"&E{0}&F{0}&"_"&TEXT(A{0},"YYYYMMDD")'.format(row)

            copy_filename_formula = f'=IF(C{row}<>"", {base_filename_formula} & "_", "")'
            sheet.cell(row=row, column=copy_filename_col_idx, value=copy_filename_formula)

            for i in range(1, 6):
                photo_col_idx = headers.index(f'사진{i}') + 1
                photo_formula = f'=IF(IFERROR(VALUE({get_column_letter(photo_count_col_idx)}{row}), 0)>={i}, HYPERLINK("images\\" & {base_filename_formula} & "_{i}.png", {base_filename_formula} & "_{i}.png"), "")'
                sheet.cell(row=row, column=photo_col_idx, value=photo_formula)

            danawa_formula = f'=IF(C{row}<>"", HYPERLINK("https://search.danawa.com/dsearch.php?query=" & {search_query}, "다나와 가격 확인"), "")'
            sheet.cell(row=row, column=danawa_col_idx, value=danawa_formula)

        apply_common_validations(sheet, headers)
        apply_final_formatting(sheet, headers)

    wb.save('CPU.xlsx')

def create_gpu_excel():
    """GPU.xlsx 파일을 생성합니다. ⭐ TDP 컬럼 추가"""
    wb = Workbook()
    wb.remove(wb.active)
    create_validation_lists_sheet(wb)

    # ⭐ TDP (W) 컬럼 추가 (28번째)
    headers = ['판매글 게시일자', '날짜 선택', '브랜드', '제조사', '시리즈', '모델 번호', '접미사', '세부 모델명', '메모리 용량'] + COMMON_HEADERS + ['TDP (W)']
    today_str = date.today().strftime("%Y-%m-%d")

    for sheet_name in SHEET_NAMES:
        sheet = wb.create_sheet(sheet_name)
        sheet.append(headers)

        dv_date = DataValidation(type="list", formula1='=Date_Selection', allow_blank=True)
        dv_date.errorStyle = 'warning'
        dv_date.promptTitle = '입력 안내'
        dv_date.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_date.showInputMessage = True
        dv_date.add('B2:B1000')
        sheet.add_data_validation(dv_date)

        dv_brand = DataValidation(type="list", formula1='=GPU_Brand', allow_blank=True)
        dv_brand.errorStyle = 'warning'
        dv_brand.promptTitle = '입력 안내'
        dv_brand.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_brand.showInputMessage = True
        dv_brand.add('C2:C1000')
        sheet.add_data_validation(dv_brand)

        dv_manufacturer = DataValidation(type="list", formula1='=GPU_Manufacturer', allow_blank=True)
        dv_manufacturer.errorStyle = 'warning'
        dv_manufacturer.promptTitle = '입력 안내'
        dv_manufacturer.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_manufacturer.showInputMessage = True
        dv_manufacturer.add('D2:D1000')
        sheet.add_data_validation(dv_manufacturer)

        dv_series = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(C2," ","_")&"_Series")', allow_blank=True)
        dv_series.errorStyle = 'warning'
        dv_series.promptTitle = '입력 안내'
        dv_series.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_series.showInputMessage = True
        dv_series.add('E2:E1000')
        sheet.add_data_validation(dv_series)

        dv_model = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(C2," ","_")&"_"&E2&"_Models")', allow_blank=True)
        dv_model.errorStyle = 'warning'
        dv_model.promptTitle = '입력 안내'
        dv_model.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_model.showInputMessage = True
        dv_model.add('F2:F1000')
        sheet.add_data_validation(dv_model)

        dv_suffix = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(C2," ","_")&"_Suffix")', allow_blank=True)
        dv_suffix.errorStyle = 'warning'
        dv_suffix.promptTitle = '입력 안내'
        dv_suffix.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_suffix.showInputMessage = True
        dv_suffix.add('G2:G1000')
        sheet.add_data_validation(dv_suffix)

        dv_detailed_model = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE(D2," ","_"),"(","")&"_GPU_Detailed_Models")', allow_blank=True)
        dv_detailed_model.errorStyle = 'warning'
        dv_detailed_model.promptTitle = '입력 안내'
        dv_detailed_model.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_detailed_model.showInputMessage = True
        dv_detailed_model.add('H2:H1000')
        sheet.add_data_validation(dv_detailed_model)

        dv_memory = DataValidation(type="list", formula1='=GPU_Memory_Capacity', allow_blank=True)
        dv_memory.errorStyle = 'warning'
        dv_memory.promptTitle = '입력 안내'
        dv_memory.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_memory.showInputMessage = True
        dv_memory.add('I2:I1000')
        sheet.add_data_validation(dv_memory)

        # ⭐ TDP 유효성 검사 추가
        tdp_col_idx = headers.index('TDP (W)') + 1
        dv_tdp = DataValidation(type="list", formula1='=GPU_TDP', allow_blank=True)
        dv_tdp.errorStyle = 'warning'
        dv_tdp.promptTitle = '입력 안내'
        dv_tdp.prompt = '목록에서 TDP 값을 선택하거나 직접 입력하세요 (숫자만).'
        dv_tdp.showInputMessage = True
        dv_tdp.add(f'{get_column_letter(tdp_col_idx)}2:{get_column_letter(tdp_col_idx)}1000')
        sheet.add_data_validation(dv_tdp)

        photo_count_col_idx = headers.index('사진 개수') + 1
        copy_filename_col_idx = headers.index('파일명(복사용)') + 1
        danawa_col_idx = headers.index('다나와 검색') + 1

        dv_photo_count = DataValidation(type="list", formula1='=Photo_Count', allow_blank=True)
        dv_photo_count.errorStyle = 'warning'
        dv_photo_count.promptTitle = '입력 안내'
        dv_photo_count.prompt = '목록에서 사진 개수를 선택하세요.'
        dv_photo_count.showInputMessage = True
        dv_photo_count.add(f'{get_column_letter(photo_count_col_idx)}2:{get_column_letter(photo_count_col_idx)}1000')
        sheet.add_data_validation(dv_photo_count)

        for row in range(2, 1001):
            date_formula = f'=IF(B{row}="오늘",DATEVALUE("{today_str}"), IF(ISNUMBER(FIND("일 전", B{row})), DATEVALUE("{today_str}") - VALUE(SUBSTITUTE(B{row},"일 전","")), IF(ISNUMBER(FIND("개월 전", B{row})), EDATE(DATEVALUE("{today_str}"), -VALUE(SUBSTITUTE(B{row},"개월 전",""))), "")))'
            sheet.cell(row=row, column=1, value=date_formula)

            search_query = 'D{0}&" "&C{0}&" "&E{0}&" "&F{0}&" "&G{0}&" "&H{0}&" "&I{0}'.format(row)
            base_filename_formula = 'SUBSTITUTE(D{0}," ","_")&"_"&SUBSTITUTE(C{0}," ","_")&"_"&E{0}&F{0}&IF(G{0}<>"", "_"&G{0}, "")&"_"&SUBSTITUTE(H{0}," ","_")&"_"&I{0}&"_"&TEXT(A{0},"YYYYMMDD")'.format(row)

            copy_filename_formula = f'=IF(C{row}<>"", {base_filename_formula} & "_", "")'
            sheet.cell(row=row, column=copy_filename_col_idx, value=copy_filename_formula)

            for i in range(1, 6):
                photo_col_idx = headers.index(f'사진{i}') + 1
                photo_formula = f'=IF(IFERROR(VALUE({get_column_letter(photo_count_col_idx)}{row}), 0)>={i}, HYPERLINK("images\\" & {base_filename_formula} & "_{i}.png", {base_filename_formula} & "_{i}.png"), "")'
                sheet.cell(row=row, column=photo_col_idx, value=photo_formula)

            danawa_formula = f'=IF(C{row}<>"", HYPERLINK("https://search.danawa.com/dsearch.php?query=" & {search_query}, "다나와 가격 확인"), "")'
            sheet.cell(row=row, column=danawa_col_idx, value=danawa_formula)

        apply_common_validations(sheet, headers)
        apply_final_formatting(sheet, headers)

    wb.save('GPU.xlsx')

def create_mainboard_excel():
    """Mainboard.xlsx 파일을 생성합니다. ⭐ 소켓, 메모리 타입, 폼팩터 컬럼 추가"""
    wb = Workbook()
    wb.remove(wb.active)
    create_validation_lists_sheet(wb)

    # ⭐ 소켓, 메모리 타입, 폼팩터 컬럼 추가 (27-29번째)
    headers = ['판매글 게시일자', '날짜 선택', '제조사', '시리즈', '칩셋', '모델명', '세부특징1', '세부특징2'] + COMMON_HEADERS + ['소켓', '메모리 타입', '폼팩터']
    today_str = date.today().strftime("%Y-%m-%d")

    for sheet_name in SHEET_NAMES:
        sheet = wb.create_sheet(sheet_name)
        sheet.append(headers)

        dv_date = DataValidation(type="list", formula1='=Date_Selection', allow_blank=True)
        dv_date.errorStyle = 'warning'
        dv_date.promptTitle = '입력 안내'
        dv_date.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_date.showInputMessage = True
        dv_date.add('B2:B1000')
        sheet.add_data_validation(dv_date)

        dv_manufacturer = DataValidation(type="list", formula1='=Mainboard_Manufacturer', allow_blank=True)
        dv_manufacturer.errorStyle = 'warning'
        dv_manufacturer.promptTitle = '입력 안내'
        dv_manufacturer.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_manufacturer.showInputMessage = True
        dv_manufacturer.add('C2:C1000')
        sheet.add_data_validation(dv_manufacturer)

        dv_series = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(C2," ","_")&"_Series")', allow_blank=True)
        dv_series.errorStyle = 'warning'
        dv_series.promptTitle = '입력 안내'
        dv_series.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_series.showInputMessage = True
        dv_series.add('D2:D1000')
        sheet.add_data_validation(dv_series)

        dv_chipset = DataValidation(type="list", formula1='=Mainboard_Chipset', allow_blank=True)
        dv_chipset.errorStyle = 'warning'
        dv_chipset.promptTitle = '입력 안내'
        dv_chipset.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_chipset.showInputMessage = True
        dv_chipset.add('E2:E1000')
        sheet.add_data_validation(dv_chipset)

        dv_model = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(C2," ","_")&"_"&SUBSTITUTE(SUBSTITUTE(D2," ","_"),"-","_")&"_Models")', allow_blank=True)
        dv_model.errorStyle = 'warning'
        dv_model.promptTitle = '입력 안내'
        dv_model.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_model.showInputMessage = True
        dv_model.add('F2:F1000')
        sheet.add_data_validation(dv_model)

        dv_suffix1 = DataValidation(type="list", formula1='=Mainboard_Suffix', allow_blank=True)
        dv_suffix1.errorStyle = 'warning'
        dv_suffix1.promptTitle = '입력 안내'
        dv_suffix1.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_suffix1.showInputMessage = True
        dv_suffix1.add('G2:G1000')
        sheet.add_data_validation(dv_suffix1)

        dv_suffix2 = DataValidation(type="list", formula1='=Mainboard_Suffix', allow_blank=True)
        dv_suffix2.errorStyle = 'warning'
        dv_suffix2.promptTitle = '입력 안내'
        dv_suffix2.prompt = '목록에서 선택하거나 직접 입력할 수 있습니다.'
        dv_suffix2.showInputMessage = True
        dv_suffix2.add('H2:H1000')
        sheet.add_data_validation(dv_suffix2)

        # ⭐ 소켓, 메모리 타입, 폼팩터 유효성 검사 추가
        socket_col_idx = headers.index('소켓') + 1
        dv_socket = DataValidation(type="list", formula1='=Mainboard_Socket', allow_blank=True)
        dv_socket.errorStyle = 'warning'
        dv_socket.promptTitle = '입력 안내'
        dv_socket.prompt = '목록에서 소켓을 선택하세요.'
        dv_socket.showInputMessage = True
        dv_socket.add(f'{get_column_letter(socket_col_idx)}2:{get_column_letter(socket_col_idx)}1000')
        sheet.add_data_validation(dv_socket)

        memory_type_col_idx = headers.index('메모리 타입') + 1
        dv_memory_type = DataValidation(type="list", formula1='=Mainboard_Memory_Type', allow_blank=True)
        dv_memory_type.errorStyle = 'warning'
        dv_memory_type.promptTitle = '입력 안내'
        dv_memory_type.prompt = '목록에서 메모리 타입을 선택하세요.'
        dv_memory_type.showInputMessage = True
        dv_memory_type.add(f'{get_column_letter(memory_type_col_idx)}2:{get_column_letter(memory_type_col_idx)}1000')
        sheet.add_data_validation(dv_memory_type)

        form_factor_col_idx = headers.index('폼팩터') + 1
        dv_form_factor = DataValidation(type="list", formula1='=Mainboard_Form_Factor', allow_blank=True)
        dv_form_factor.errorStyle = 'warning'
        dv_form_factor.promptTitle = '입력 안내'
        dv_form_factor.prompt = '목록에서 폼팩터를 선택하세요.'
        dv_form_factor.showInputMessage = True
        dv_form_factor.add(f'{get_column_letter(form_factor_col_idx)}2:{get_column_letter(form_factor_col_idx)}1000')
        sheet.add_data_validation(dv_form_factor)

        photo_count_col_idx = headers.index('사진 개수') + 1
        copy_filename_col_idx = headers.index('파일명(복사용)') + 1
        danawa_col_idx = headers.index('다나와 검색') + 1

        dv_photo_count = DataValidation(type="list", formula1='=Photo_Count', allow_blank=True)
        dv_photo_count.errorStyle = 'warning'
        dv_photo_count.promptTitle = '입력 안내'
        dv_photo_count.prompt = '목록에서 사진 개수를 선택하세요.'
        dv_photo_count.showInputMessage = True
        dv_photo_count.add(f'{get_column_letter(photo_count_col_idx)}2:{get_column_letter(photo_count_col_idx)}1000')
        sheet.add_data_validation(dv_photo_count)

        for row in range(2, 1001):
            date_formula = f'=IF(B{row}="오늘",DATEVALUE("{today_str}"), IF(ISNUMBER(FIND("일 전", B{row})), DATEVALUE("{today_str}") - VALUE(SUBSTITUTE(B{row},"일 전","")), IF(ISNUMBER(FIND("개월 전", B{row})), EDATE(DATEVALUE("{today_str}"), -VALUE(SUBSTITUTE(B{row},"개월 전",""))), "")))'
            sheet.cell(row=row, column=1, value=date_formula)

            search_query = 'C{0}&" "&D{0}&" "&E{0}&" "&F{0}&" "&G{0}&" "&H{0}'.format(row)
            base_filename_formula = 'SUBSTITUTE(C{0}," ","_")&"_"&SUBSTITUTE(D{0}," ","_")&"_"&E{0}&"_"&SUBSTITUTE(F{0}," ","_")&IF(G{0}<>"", "_"&SUBSTITUTE(G{0},"-","_"), "")&IF(H{0}<>"", "_"&SUBSTITUTE(H{0},"-","_"), "")&"_"&TEXT(A{0},"YYYYMMDD")'.format(row)

            copy_filename_formula = f'=IF(C{row}<>"", {base_filename_formula} & "_", "")'
            sheet.cell(row=row, column=copy_filename_col_idx, value=copy_filename_formula)

            for i in range(1, 6):
                photo_col_idx = headers.index(f'사진{i}') + 1
                photo_formula = f'=IF(IFERROR(VALUE({get_column_letter(photo_count_col_idx)}{row}), 0)>={i}, HYPERLINK("images\\" & {base_filename_formula} & "_{i}.png", {base_filename_formula} & "_{i}.png"), "")'
                sheet.cell(row=row, column=photo_col_idx, value=photo_formula)

            danawa_formula = f'=IF(C{row}<>"", HYPERLINK("https://search.danawa.com/dsearch.php?query=" & {search_query}, "다나와 가격 확인"), "")'
            sheet.cell(row=row, column=danawa_col_idx, value=danawa_formula)

        apply_common_validations(sheet, headers)
        apply_final_formatting(sheet, headers)

    wb.save('Mainboard.xlsx')


# --- 메인 실행 ---
if __name__ == "__main__":
    # 스크립트가 실행되는 디렉토리로 작업 경로를 변경합니다.
    # 이렇게 하면 항상 스크립트 파일 위치에 엑셀 파일이 생성됩니다.
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)

    try:
        print("CPU.xlsx 생성 중... (소켓 컬럼 포함)")
        create_cpu_excel()
        print("✅ CPU.xlsx 생성 완료!")

        print("GPU.xlsx 생성 중... (TDP 컬럼 포함)")
        create_gpu_excel()
        print("✅ GPU.xlsx 생성 완료!")

        print("Mainboard.xlsx 생성 중... (소켓, 메모리 타입, 폼팩터 컬럼 포함)")
        create_mainboard_excel()
        print("✅ Mainboard.xlsx 생성 완료!")

        print("\n🎉 모든 엑셀 파일이 성공적으로 생성되었습니다!")
        print("📍 생성 위치: 스크립트 파일과 동일한 디렉토리")
        print("\n✅ 추가된 컬럼:")
        print("   - CPU.xlsx: 소켓")
        print("   - GPU.xlsx: TDP (W)")
        print("   - Mainboard.xlsx: 소켓, 메모리 타입, 폼팩터")
    except Exception as e:
        print(f"❌ 오류가 발생했습니다: {e}")
        print("열려있는 엑셀 파일을 모두 닫고 다시 시도해보세요.")
        print("openpyxl 라이브러리가 설치되었는지 확인해주세요. (터미널에서 pip install openpyxl 실행)")
