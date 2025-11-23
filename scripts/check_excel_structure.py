#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 파일 구조 확인 스크립트
"""

import openpyxl
import json

def analyze_excel(file_path):
    """엑셀 파일의 구조를 분석합니다."""
    print(f"\n📊 파일 분석: {file_path}")
    print("=" * 60)

    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n📋 시트: {sheet_name}")
        print("-" * 60)

        # 헤더 (첫 번째 행)
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(cell.value)

        print(f"컬럼 개수: {len(headers)}")
        print(f"컬럼 목록:")
        for i, h in enumerate(headers, 1):
            print(f"  {i:2d}. {h}")

        # 데이터 개수
        row_count = 0
        for row in sheet.iter_rows(min_row=2):
            if any(cell.value for cell in row):
                row_count += 1
        print(f"\n데이터 행 개수: {row_count}")

        # 샘플 데이터 (첫 2개)
        if row_count > 0:
            print(f"\n샘플 데이터:")
            for row_idx in range(2, min(4, row_count + 2)):
                print(f"\n  [행 {row_idx}]")
                for col_idx, header in enumerate(headers, 1):
                    value = sheet.cell(row=row_idx, column=col_idx).value
                    if value:
                        print(f"    {header}: {value}")

if __name__ == "__main__":
    files = [
        "../datas/CPU.xlsx",
        "../datas/GPU.xlsx",
        "../datas/Mainboard.xlsx"
    ]

    for file in files:
        try:
            analyze_excel(file)
        except Exception as e:
            print(f"❌ {file} 읽기 실패: {e}")

    print("\n" + "=" * 60)
    print("✅ 분석 완료!")
